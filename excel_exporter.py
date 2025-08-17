import pandas as pd
import io
from typing import List, Dict, Any
import streamlit as st
from datetime import datetime
from openpyxl.styles import NamedStyle, Font, PatternFill
import logging

class ExcelExporter:
    """Enhanced Excel export with better formatting and error handling"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def export_candidates(self, candidates_data: List[Dict[str, Any]]) -> bytes:
        """
        Export candidates data to Excel file with enhanced formatting
        
        Args:
            candidates_data: List of candidate information dictionaries
            
        Returns:
            Excel file as bytes
        """
        try:
            if not candidates_data:
                raise ValueError("No candidate data to export")
            
            self.logger.info(f"Exporting {len(candidates_data)} candidates to Excel")
            
            # Prepare data for DataFrame
            excel_data = []
            
            for i, candidate in enumerate(candidates_data, 1):
                row = {
                    'Sr. No.': i,
                    'First Name': candidate.get('first_name', ''),
                    'Last Name': candidate.get('last_name', ''),
                    'Mobile': candidate.get('mobile', ''),
                    'Email': candidate.get('email', ''),
                    'Current Job Title': candidate.get('current_job_title', ''),
                    'Current Company': candidate.get('current_company', ''),
                    'Previous Job Title': candidate.get('previous_job_title', ''),
                    'Previous Company': candidate.get('previous_company', ''),
                    'Source File': candidate.get('filename', '')
                }
                excel_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Create Excel writer object
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Resume Data', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Resume Data']
                
                # Create styles
                self._apply_excel_formatting(workbook, worksheet, df)
                
                # Add summary sheet
                self._add_summary_sheet(writer, candidates_data)
            
            output.seek(0)
            self.logger.info("Excel export completed successfully")
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
            st.error(f"❌ Error creating Excel file: {str(e)}")
            raise e
    
    def _apply_excel_formatting(self, workbook, worksheet, df):
        """Apply enhanced formatting to Excel worksheet"""
        try:
            # Create header style
            header_style = NamedStyle(name="header_style")
            header_style.font = Font(bold=True, color="FFFFFF")
            header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Register the style to the workbook
            if "header_style" not in workbook.named_styles:
                workbook.add_named_style(header_style)

            # Apply style to header row
            for cell in worksheet[1]:
                cell.style = header_style
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width with padding, but cap at reasonable maximum
                adjusted_width = min(max_length + 3, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders and alternate row colors
            from openpyxl.styles import Border, Side
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to all cells with data
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, 
                                         min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
                    
        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {str(e)}")
    
    def _add_summary_sheet(self, writer, candidates_data):
        """Add summary statistics sheet"""
        try:
            # Calculate summary statistics
            total_candidates = len(candidates_data)
            candidates_with_email = len([c for c in candidates_data if c.get('email')])
            candidates_with_mobile = len([c for c in candidates_data if c.get('mobile')])
            candidates_with_current_job = len([c for c in candidates_data if c.get('current_job_title')])
            candidates_with_previous_job = len([c for c in candidates_data if c.get('previous_job_title')])
            
            # Create summary data
            summary_data = [
                ['Metric', 'Count', 'Percentage'],
                ['Total Candidates', total_candidates, '100%'],
                ['With Email', candidates_with_email, f'{(candidates_with_email/total_candidates*100):.1f}%'],
                ['With Mobile', candidates_with_mobile, f'{(candidates_with_mobile/total_candidates*100):.1f}%'],
                ['With Current Job', candidates_with_current_job, f'{(candidates_with_current_job/total_candidates*100):.1f}%'],
                ['With Previous Job', candidates_with_previous_job, f'{(candidates_with_previous_job/total_candidates*100):.1f}%'],
                ['', '', ''],
                ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
                ['Total Files Processed', total_candidates, '']
            ]
            
            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            
            # Write to Excel
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_worksheet = writer.sheets['Summary']
            
            # Apply header formatting
            for cell in summary_worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust column widths
            for column in summary_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                summary_worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Could not create summary sheet: {str(e)}")
